#!/usr/bin/env python

import openpyxl
import re
import pprint
pp = pprint.PrettyPrinter(indent=4)
from datetime import datetime, date
from tomark import Tomark

def get_datestr_if_date(v):
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return v

def get_monetary_quantity(v):
    if isinstance(v, int):
        return f"€{row[7].value}"
    return v

def make_md_links(str):
    # add no-break space in tool field
    str = str.replace("Okapi plugin", "Okapi plugin", 1)
    # align titles to left
    str = str.replace("|-----", "|:----")
    # convert dict to md link
    pattern = r"\{'(#\d+)': '(https?:.+)'\}"
    replace = r"[\1](\2)"
    return re.sub(pattern, replace, str)

file = "OmegaT_issues_nopw.xlsx"

wb = openpyxl.load_workbook(file)
ws = wb['2022']

ticket_numbers = [row[1].value for row in ws.iter_rows()
    if row[1].value != None and re.match("#\d+", row[1].value)]
ticket_links = [row[1].hyperlink.target for row in ws.iter_rows()
    if row[1].value != None and re.match("#\d+", row[1].value)]

ticket_links_dict = dict(zip(ticket_numbers, ticket_links))

#print(f"{ticket_links_dict=}")

tickets_data = []
for row in ws.iter_rows():
    if row[1].value != None and re.match("#\d+", row[1].value):
        row_dict = {
            "Created": get_datestr_if_date(row[0].value),
            "#": {row[1].value: row[1].hyperlink.target},
            "Tool": row[2].value,
            "Title": row[3].value,
            "Type": row[4].value,
            "Status": row[5].value,
            "Developer": row[6].value,
            "Cost": get_monetary_quantity(row[7].value),
            "Comments": row[8].value
        }
        tickets_data.append(row_dict)

#printable_data = [for row in tickets_data]
#squared = list(map(lambda x: x["#"] = [x["Number"](x["url"])], tickets_data))

private_keys = ['Developer', 'Cost', 'Comments']
public_data = [
    dict((key, ticket_row[key]) for key in ticket_row if key not in private_keys)
    for ticket_row in tickets_data
    ]

md_tbl = "# OmegaT and Okapi tickets opened (or still open) since 2020\n\n" + Tomark.table(public_data)
md_tbl = make_md_links(md_tbl)

with open('omegat-tickets.md', 'w') as f:
    f.write(md_tbl)
